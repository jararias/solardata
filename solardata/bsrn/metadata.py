
from __future__ import absolute_import, print_function, division

import os
import json
import datetime
import warnings
import itertools as itt
from copy import copy

from loguru import logger
import openpyxl as oxl
import numpy as np
import pandas as pd

from . import load_data as load_bsrn_data
from . import sites as get_bsrn_sites
from . import config
from .description_tables import TableA4, TableA5


logger.enable(__name__)


class ExcelWriter(object):
    def __init__(self):
        self.wb = oxl.Workbook()
        self._sheet_counter = 0
        self.ws = None
        self.cur_row = None
        self.cur_column = None

    def add_sheet(self, name=None):
        if self._sheet_counter == 0:
            self.ws = self.wb.active
        else:
            self.ws = self.wb.create_sheet(name)
        self._sheet_counter += 1
        name = f'Sheet_{self._sheet_counter}' if name is None else name
        self.ws.title = name
        self.cur_row = 1
        self.cur_column = 1

    def write(self, text, row=None, column=None, ha='left'):
        cur_row = row or self.cur_row
        cur_column = column or self.cur_column
        c = self.ws.cell(row=cur_row, column=cur_column, value=text)
        c.alignment = oxl.styles.Alignment(horizontal=ha)
        self.cur_row = cur_row + 1
        self.cur_column = cur_column
        return c

    def scale_column_width(self, column_number, scale_factor):
        column_name = oxl.utils.get_column_letter(column_number)
        column_width = self.ws.column_dimensions[column_name].width
        self.ws.column_dimensions[column_name].width = (
            column_width * scale_factor)

    def set_column_bestfit(self):
        for column in range(1, self.ws.max_column+1):
            column_name = oxl.utils.get_column_letter(column)
            self.ws.column_dimensions[column_name].bestFit = True

    def save(self, fname):
        self.wb.save(fname)


class BSRNMetadata(object):
    def __init__(self, site):
        if site not in get_bsrn_sites():
            raise ValueError(f'unknown BSRN site `{site}`')
        self.site = site
        init_year = 1992
        last_year = datetime.datetime.now().year
        self._years = range(init_year, last_year)
        local_path = config.load().get('localdir')
        self.fname = os.path.join(local_path, site, f'metadata_{site}.json')
        self.metadata = {}
        if os.path.exists(self.fname):
            self.metadata = json.load(open(self.fname, 'rt'))

    def __enter__(self):
        return self

    def __exit__(self, *args, **kwargs):  # type, value, traceback):
        self.save()
        return True

    def save(self, fname=None):
        json.dump(self.metadata, open(fname or self.fname, 'wt'))

    def get(self, year, month, force_reload=False):
        period = f'{year}{month:02d}'
        do_reload = True
        if period in self.metadata:
            do_reload = (len(self.metadata[period]) == 0) and (force_reload is True)
        if do_reload:
            logger.debug(f'Retrieving year={year} & month={month} @ {self.site}')

            result = load_bsrn_data(
                self.site, year, month, full_output=True,
                check_remote_server_on_missing_file=False)

            if result is None:
                logger.warning(f'missing period {period}')
                self.metadata[period] = {}
                return self.metadata[period]

            _, metadata, logrec = result

            if 'surface_type' in metadata:
                metadata['surface_type'] = TableA4.get(
                    metadata['surface_type'], metadata['surface_type'])

            if 'topograpy_type' in metadata:
                metadata['topography_type'] = metadata.pop('topograpy_type')

            if 'topography_type' in metadata:
                metadata['topography_type'] = TableA5.get(
                    metadata['topography_type'], metadata['topography_type'])

            availability = {}
            for lrid, lrdata in logrec.items():
                availability[lrid] = {}
                for key in filter(lambda k: k not in ('description', 'utc_times'), lrdata):
                    values = lrdata[key]
                    if (n_samples := len(lrdata['utc_times'])) > 0:
                        n_valid_samples = len(values[~np.isnan(values)])
                        availability[lrid][key] = n_valid_samples / n_samples
                availability[lrid]['description'] = lrdata['description']
            metadata['data_availability'] = availability

            self.metadata[period] = metadata

        return self.metadata[period]

    def get_chronology(self, what=None, years=None, force_reload=False):
        chronology = {}
        for year, month in itt.product(years or self._years, range(1, 13)):
            if (metadata := self.get(year, month, force_reload=force_reload)) is None:
                continue
            chronology[f'{year}{month:02d}'] = (
                metadata if what is None else metadata.get(what, None))
        return chronology

    def get_messages(self, years=None):
        chronology = self.get_chronology(years=years)
        return {period: chronology[period].get('message', '')
                for period in chronology}

    def get_instruments(self, instrument_id=None, years=None):
        chronology = self.get_chronology(years=years)

        instr_chrono = {}
        for period, metadata in chronology.items():

            if 'measurements' not in metadata:
                logger.warning(f'missing `measurements` @ {self.site}:{period}')
                continue

            for measurement in metadata['measurements']:
                instr_id = str(measurement['id_instr'])
                id_qty = measurement.get('id_qty', None)
                qties = metadata['quantity_measured']

                instr_chrono.setdefault(instr_id, {})
                instr_chrono[instr_id].setdefault(period, {})
                instr_chrono[instr_id][period].update(
                    {'quantity_measured': qties.get(id_qty, qties.get(str(id_qty), {}))}
                )

                instr_data = None
                for instrument in metadata['instruments']:
                    if instr_id == str(instrument['wrmc_id']):
                        instr_data = instrument
                        break

                if instr_data is None:
                    warnings.warn(f'missing instrument {instr_id} for {period}', UserWarning)
                    continue

                instr_chrono[instr_id][period].update(
                    {
                        'manufacturer': instr_data.get('manufacturer'),
                        'model': instr_data.get('model'),
                        'serial_number': instr_data.get('serial_number'),
                        'wrmc_id': str(instr_data.get('wrmc_id'))
                    }
                )

                calibration_bands = {}
                for band in range(1, 4):
                    calibration_bands[f'calibration_band_{band}'] = {
                        kw: instr_data.get(f'{kw}_of_band_{band}')
                        for kw in ('wavelength', 'bandwidth', 'start_of_calibration_period',
                                   'end_of_calibration_period', 'mean_calibration_coefficient',
                                   'standard_error_of_calibration_coefficient', 'number_of_comparisons')
                    }
                instr_chrono[instr_id][period].update(calibration_bands)

                for attr_name in filter(lambda name: '_band_' not in name, instr_data):
                    instr_chrono[instr_id][period].setdefault(attr_name, instr_data.get(attr_name))

        # remove void calibration bands

        def missing_period(instr_id, period, band_number):
            band_name = f'calibration_band_{band_number}'
            this_instr_chrono = instr_chrono[instr_id][period]
            if (instr_band := this_instr_chrono.get(band_name, None)) is None:
                return True
            return all(val is None or val == 'XXX' for val in instr_band.values())

        for instr_id, band_number in itt.product(instr_chrono, range(1, 4)):
            periods = instr_chrono[instr_id]
            band_name = f'calibration_band_{band_number}'
            if all(missing_period(instr_id, period, band_number) for period in periods):
                for period in instr_chrono[instr_id]:
                    instr_chrono[instr_id][period] = {
                        k: v for k, v in instr_chrono[instr_id][period].items()
                        if k != band_name
                    }

        if instrument_id is None:
            return instr_chrono

        return instr_chrono.get(instrument_id, None)

    def get_calibration_record(self, instrument_id, years=None):
        instrument = self.get_instruments(instrument_id, years=years)

        data = {1: [], 2: [], 3: []}
        for period, ivalues in instrument.items():
            for band_number in range(1, 4):
                band_name = f'calibration_band_{band_number}'
                if band_name in ivalues:
                    coef_value = ivalues[band_name].get(
                        'mean_calibration_coefficient', np.nan)
                    data[band_number].append((period, coef_value))

        for band_number in data:
            if len(data[band_number]) == 0:
                continue
            times = pd.to_datetime(
                [v[0] for v in data[band_number]], format='%Y%m')
            coefs = np.array([v[1] for v in data[band_number]])
            data[band_number] = pd.Series(data=coefs, index=times, copy=True)

        df = pd.DataFrame({
            f'band{k}': v for k, v in data.items() if len(v) > 0})
        return df if len(df) > 0 else None

    def to_excel(self, excel_fname, years=None):
        chronology = self.get_chronology(years=years)
        chronology = {k: v for k, v in chronology.items() if v}
        instruments = self.get_instruments(years=years)

        excel = ExcelWriter()

        # INSTRUMENTS & CALIBRATION

        for instr_id, instr_chrono in instruments.items():
            excel.add_sheet(name=f'Instrument {instr_id}')
            logger.info(f'Writing worksheet `Instrument {instr_id}`')

            cell = excel.write('WVLn: wavelength of calibration band n',
                               row=1, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write('WBDn: waveband of calibration band n',
                               row=2, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write('STRn: start of calibration period of band n',
                               row=3, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write('ENDn: end of calibration period of band n',
                               row=4, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write(
                'COEn: mean of calibration coefficient of band n',
                row=5, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write(
                'STDn: standard error of calibration coefficient of band n',
                row=6, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)
            cell = excel.write('CMPn: number of comparisons of band n',
                               row=7, column=1, ha='left')
            cell.font = oxl.styles.Font(b=True)

            header_row = 9
            for n_column, hdr_name in enumerate(
                    ('YearMonth', 'Manufacturer', 'Model', 'Serial Number',
                     'WRMC ID', 'Description', 'Unit')):
                column = n_column + 1
                cell = excel.write(hdr_name, row=header_row, column=column)
                cell.font = oxl.styles.Font(b=True)

            for n_period, period in enumerate(instr_chrono):
                attributes = copy(instr_chrono[period])

                row = n_period + 10
                column = 1

                excel.write(period, row=row, column=column)
                column += 1

                def write_instr(attr_name):
                    # pylint: disable=cell-var-from-loop
                    text = attributes.pop(attr_name, '') or ''
                    excel.write(text, row=row, column=column)
                    return column + 1

                column = write_instr('manufacturer')
                column = write_instr('model')
                column = write_instr('serial_number')
                column = write_instr('wrmc_id')
                column = write_instr('description')
                column = write_instr('unit')

                # calibration data...
                for band_number in range(1, 4):

                    if f'calibration_band_{band_number}' in attributes:
                        band = attributes[f'calibration_band_{band_number}']

                        def write_band(attr_name, is_date=False):
                            # pylint: disable=cell-var-from-loop
                            text = band.get(attr_name, '') or ''
                            cell = excel.write(text, row=row, column=column)
                            if (is_date is True) and (text != ''):
                                try:
                                    month, day, year = [
                                        int(v) for v in text.split('/')]
                                    year += 1900 if year > 50 else 2000
                                    cell.value = datetime.date(
                                        year, month, day)
                                    cell.number_format = 'yyyy/mm/dd'
                                except Exception:  # pylint: disable=broad-except
                                    pass
                            return column + 1

                        start_column = column
                        column = write_band('wavelength')
                        column = write_band('waveband')
                        column = write_band('start_of_calibration_period', True)
                        column = write_band('end_of_calibration_period', True)
                        column = write_band('mean_calibration_coefficient')
                        column = write_band(
                            'standard_error_of_calibration_coefficient')
                        column = write_band('number_of_comparisons')

                        if n_period == 0:

                            def write_period_0(txt):
                                # pylint: disable=cell-var-from-loop
                                cell = excel.write(
                                    txt, row=header_row, column=column)
                                cell.font = oxl.styles.Font(b=True)
                                return column + 1

                            column = start_column
                            column = write_period_0(f'WVL{band_number}')
                            column = write_period_0(f'WBD{band_number}')
                            column = write_period_0(f'STR{band_number}')
                            column = write_period_0(f'END{band_number}')
                            column = write_period_0(f'COE{band_number}')
                            column = write_period_0(f'STD{band_number}')
                            column = write_period_0(f'CMP{band_number}')

                        attributes = {
                            k: v for k, v in attributes.items()
                            if k != f'calibration_band_{band_number}'}

                # remaining attributes...
                for attr_name in attributes:

                    attr_value = attributes.get(attr_name, '')
                    if attr_value is None:
                        attr_value = ''
                    if not isinstance(attr_value, str):
                        attr_value = repr(attr_value)
                    cell = excel.write(attr_value, row=row, column=column)

                    if attr_name == 'calibration_remarks':
                        attr_value = ', '.join(attributes.get(attr_name))
                        cell.value = attr_value

                    if attr_name == 'changed_on':
                        attr_value = attributes.get(attr_name)
                        if any([v is None for v in attr_value.values()]):
                            cell.value = ''
                        else:
                            try:
                                cell.value = datetime.datetime(
                                    int(period[:4]), int(period[4:]),
                                    attr_value['day'], attr_value['hour'],
                                    attr_value['minute'])
                                cell.number_format = 'yyyy/mm/dd hh:mm'
                            except Exception:  # pylint: disable=broad-except
                                pass

                    if n_period == 0:
                        cell = excel.write(
                            attr_name, row=header_row, column=column)
                        cell.font = oxl.styles.Font(b=True)

                    column = column + 1

            excel.set_column_bestfit()

        # MESSAGES

        logger.info('Writing worksheet `Messages`')
        messages = self.get_messages(years=years)
        excel.add_sheet(name='Messages')

        cell = excel.write('Period', row=1, column=1, ha='left')
        cell.font = oxl.styles.Font(b=True)
        cell = excel.write('Message', row=1, column=2, ha='left')
        cell.font = oxl.styles.Font(b=True)

        for n_message, (period, message) in enumerate(messages.items()):
            excel.write(period, row=n_message+2, column=1, ha='left')
            excel.write(message, row=n_message+2, column=2, ha='left')

        excel.set_column_bestfit()
        excel.wb._sheets.insert(0, excel.ws)  # pylint: disable=protected-access
        excel.wb._sheets.pop(-1)  # pylint: disable=protected-access

        # GENERAL INFO

        logger.info('Writing worksheet `Basic info`')
        excel.add_sheet(name='Basic info')

        cell = excel.write('Period', row=1, column=1, ha='left')
        cell.font = oxl.styles.Font(b=True)

        for n_period, period in enumerate(chronology):

            def write(attr_name):
                attr_value = chronology[period].get(attr_name, '') or ''
                cell = excel.write(attr_value, row=n_period+2, column=column)
                if attr_name in ('latitude', 'longitude'):
                    cell.number_format = '#,####0.0000'
                if n_period == 0:
                    cell = excel.write(attr_name, row=1, column=column)
                    cell.font = oxl.styles.Font(b=True)
                return column + 1

            excel.write(period, row=n_period+2, column=1)

            column = 2
            column = write('longitude')
            column = write('latitude')
            column = write('altitude')
            column = write('surface_type')
            column = write('topography_type')
            column = write('synop_id')
            column = write('data_version')
            column = write('deputy_name')
            column = write('deputy_address')
            column = write('deputy_email')
            column = write('scientist_name')
            column = write('scientist_address')
            column = write('scientist_email')
            column = write('station_address')
            column = write('station_email')
            column = write('station_id')

        excel.set_column_bestfit()
        excel.wb._sheets.insert(0, excel.ws)  # pylint: disable=protected-access
        excel.wb._sheets.pop(-1)  # pylint: disable=protected-access

        # DATA AVAILABILITY

        lrid = '0100'
        availability = {
            period: chrono.get('data_availability', {}).get(lrid, {})
            for period, chrono in chronology.items()}

        logger.info('Writing worksheet `Availability LR0100`')
        excel.add_sheet(name='Availability LR0100')

        periods = list(availability.keys())
        description = availability[periods[0]].get('description', 'unknown')
        cell = excel.write(f'Logical record {lrid}: {description}',
                           row=1, column=1, ha='left')
        cell.font = oxl.styles.Font(b=True)

        variables = []
        for values in availability.values():
            variables.extend(list(values.keys()))
        variables = sorted(set(variables))
        if 'description' in variables:
            variables.pop(variables.index('description'))

        cell = excel.write('Period', row=2, column=1, ha='left')
        cell.font = oxl.styles.Font(b=True)
        for n_variable, variable in enumerate(variables):
            column = n_variable+2
            cell = excel.write(variable, row=2, column=column, ha='right')
            cell.font = oxl.styles.Font(b=True)

        for n_period, (period, avail) in enumerate(availability.items()):
            row = n_period+3
            excel.write(period, row=n_period+3, column=1, ha='left')
            for n_variable, variable in enumerate(variables):
                column = n_variable+2
                value = avail.get(variable, '')
                cell = excel.write(value, row=row, column=column, ha='right')
                cell.number_format = '#,##0.00'

        excel.set_column_bestfit()

        excel.save(excel_fname)


def get_chronology(site, years=None, what=None, force_reload=False):
    with BSRNMetadata(site) as mtdt:
        chronology = mtdt.get_chronology(
            years=years, what=what, force_reload=force_reload)
    return chronology


def get_instruments_chronology(site, years=None):
    with BSRNMetadata(site) as mtdt:
        instruments = mtdt.get_instruments(years=years)
    return instruments


def get_instruments_description(site, years=None):
    properties = ('manufacturer', 'model', 'wrmc_id', 'serial_number')
    instruments = get_instruments_chronology(site, years)
    description = {}
    for iid in instruments:
        e0 = instruments[iid][list(instruments[iid].keys())[0]]
        description[iid] = {k: e0[k] for k in properties}
    return description


def get_calibration_record(site, instrument_id, years=None):
    with BSRNMetadata(site) as mtdt:
        data = mtdt.get_calibration_record(instrument_id, years)
    return data


def to_excel(site, excel_fname, years=None):
    dirname = os.path.dirname(os.path.abspath(excel_fname))
    if not os.path.exists(dirname):
        os.makedirs(dirname)
    with BSRNMetadata(site) as mtdt:
        mtdt.to_excel(excel_fname, years=years)


# if __name__ == '__main__':

#     site = 'pal'
#     chronology = get_chronology(site)
#     instruments = get_instruments_description(site)
#     calibration = {iid: get_calibration_record(site, iid)
#                    for iid in instruments}
#     to_excel(site, f'{site}_chronology.xlsx')

#     import pylab as pl
#     pl.subplot(111)
#     for iid in calibration:
#         if calibration[iid] is None:
#             continue
#         instr = instruments[iid]
#         label = f'{iid}[{instr["manufacturer"]}-{instr["model"]}]'
#         times = calibration[iid].index
#         pl.plot_date(times, calibration[iid]['band1'], label=label)
#     pl.legend()
#     pl.show()
